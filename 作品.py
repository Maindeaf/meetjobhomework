# -*- coding: utf-8 -*-
"""
Created on Sun Jul 10 10:56:02 2022

@author: user
"""

import openpyxl as op#匯入所需工具
import os
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles import Border,Side
import datetime
from datetime import datetime
from datetime import time
from datetime import timedelta
def convert(seconds):
    hour = seconds // 3600
    seconds %= 3600
    minutes = seconds // 60
    seconds %= 60
    return "%04d:%02d:%02d" % (hour, minutes, seconds)
#########################
gg = op.load_workbook(r"")#
#########################
Path = os.path.join('')#
bomlist = []
for excel in os.listdir(Path):
    bomlist.append(op.load_workbook(Path+'\\'+excel))
#########################
print(bomlist[5].worksheets[0])#測試
print(bomlist[5].worksheets[1])
#########################
gg.worksheets[0].delete_cols(1,3)#完成格式
gg.worksheets[0].delete_cols(3)
gg.worksheets[0].delete_cols(5,5)
gg.worksheets[0].delete_cols(9)
gg.worksheets[0].delete_cols(11,2)
gg.worksheets[0].delete_cols(16)
gg.worksheets[0].insert_cols(15,5)
gg.worksheets[0]['O4'].value = "平均人時(1PCS)"
gg.worksheets[0]['P4'].value = "平均機時(1PCS)"
gg.worksheets[0]['Q4'].value = "本次標準人時"
gg.worksheets[0]['R4'].value = "本次標準機時"
gg.worksheets[0]['S4'].value = "正負差"
font = Font(name = "新細明體", size=8, bold = True ,color = "FF0000")
red = Font(name = "新細明體", size=10, bold = False ,color = "FF0000")
black = Font(name = "新細明體", size=10, bold = False ,color = "000000")
grayfill = PatternFill(fill_type = 'solid',start_color='D9D9D9',end_color='D9D9D9')
gg.worksheets[0]['O4'].font = font
gg.worksheets[0]['P4'].font = font
gg.worksheets[0]['Q4'].font = font
gg.worksheets[0]['R4'].font = font
gg.worksheets[0]['S4'].font = font
gg.worksheets[0].column_dimensions['A'].hidden= True
#########################
for bom in bomlist:#匯入時間
    for i in range(4,len(bom.worksheets[1]['A'])+1):
        for j in range(5,len(gg.worksheets[0]['B'])+1):
            if (gg.worksheets[0]['C'+str(j)].value) != None:
                if (bom.worksheets[0]['A4'].value) in gg.worksheets[0]['C'+str(j)].value:
                    if (bom.worksheets[1]['A'+str(i)].value) in (gg.worksheets[0]['B'+str(j)].value):
                        if (bom.worksheets[1]['C'+str(i)].value) in (gg.worksheets[0]['H'+str(j)].value):
                            if (bom.worksheets[1]['M'+str(i)].value) != None:
                                if ('R') in (bom.worksheets[1]['M'+str(i)].value) and (bom.worksheets[1]['I'+str(i)].value)!= "0000:00:00":
                                    gg.worksheets[0]['O'+str(j)].value = bom.worksheets[1]['I'+str(i)].value
                                    print(j)
                                    print(gg.worksheets[0]['O'+str(j)].value)
                                if ('R') in (bom.worksheets[1]['M'+str(i)].value) and (bom.worksheets[1]['K'+str(i)].value)!= "0000:00:00":
                                    gg.worksheets[0]['P'+str(j)].value = bom.worksheets[1]['K'+str(i)].value
                                    print(j)
                                    print(gg.worksheets[0]['P'+str(j)].value)
#########################
for i in range(5,len(gg.worksheets[0]['B'])+1):#時間與數量計算
    if gg.worksheets[0]['O'+str(i)].value != None:
        gg.worksheets[0]['Q'+str(i)].value = convert(int(gg.worksheets[0]['I'+str(i)].value*int(timedelta(
            hours = int(gg.worksheets[0]['O'+str(i)].value[0:4]),
            minutes = int(gg.worksheets[0]['O'+str(i)].value[5:7]),
            seconds = int(gg.worksheets[0]['O'+str(i)].value[8:10])
        ).total_seconds())))
    if gg.worksheets[0]['P'+str(i)].value != None:
        gg.worksheets[0]['R'+str(i)].value = convert(int(gg.worksheets[0]['I'+str(i)].value*int(timedelta(
            hours = int(gg.worksheets[0]['P'+str(i)].value[0:4]),
            minutes = int(gg.worksheets[0]['P'+str(i)].value[5:7]),
            seconds = int(gg.worksheets[0]['P'+str(i)].value[8:10])
        ).total_seconds())))
#########################
for i in range(5,len(gg.worksheets[0]['B'])+1):#正負差
    if gg.worksheets[0]['O'+str(i)].value != None:
        gg.worksheets[0]['S'+str(i)].value = int(timedelta(
            hours = int(gg.worksheets[0]['M'+str(i)].value[0:2]),
            minutes = int(gg.worksheets[0]['M'+str(i)].value[3:5]),
            seconds = int(gg.worksheets[0]['M'+str(i)].value[6:8])
        ).total_seconds())-int(timedelta(
            hours = int(gg.worksheets[0]['Q'+str(i)].value[0:4]),
            minutes = int(gg.worksheets[0]['Q'+str(i)].value[5:7]),
            seconds = int(gg.worksheets[0]['Q'+str(i)].value[8:10])
        ).total_seconds())
    if gg.worksheets[0]['S'+str(i)].value != None:
        if gg.worksheets[0]['S'+str(i)].value >= 0:
            gg.worksheets[0]['S'+str(i)].font = red
        elif gg.worksheets[0]['S'+str(i)].value <= 0:
            gg.worksheets[0]['S'+str(i)].font = black
        gg.worksheets[0]['S'+str(i)].value = convert(abs(int(gg.worksheets[0]['S'+str(i)].value)))
#########################
for i in range(0,len(gg.worksheets[0][4])):
    if gg.worksheets[0][4][i].value != None:
        gg.worksheets[0][4][i].fill = grayfill
gg.worksheets[0].column_dimensions['B'].width = 22.7109375
gg.worksheets[0].column_dimensions['C'].width = 17.7109375
gg.worksheets[0].column_dimensions['D'].width = 29.7109375
gg.worksheets[0].column_dimensions['E'].width = 5.7109375
gg.worksheets[0].column_dimensions['F'].width = 8.7109375
gg.worksheets[0].column_dimensions['G'].width = 5.7109375
gg.worksheets[0].column_dimensions['H'].width = 21.7109375
gg.worksheets[0].column_dimensions['I'].width = 6.7109375
gg.worksheets[0].column_dimensions['J'].width = 4.7109375
gg.worksheets[0].column_dimensions['K'].width = 7.7109375
gg.worksheets[0].column_dimensions['L'].width = 6.7109375
gg.worksheets[0].column_dimensions['M'].width = 8.7109375
gg.worksheets[0].column_dimensions['N'].width = 13.0
gg.worksheets[0].column_dimensions['O'].width = 8.7109375
gg.worksheets[0].column_dimensions['P'].width = 13.0
gg.worksheets[0].column_dimensions['Q'].width = 13.0
gg.worksheets[0].column_dimensions['R'].width = 13.0
gg.worksheets[0].column_dimensions['S'].width = 9.7109375
gg.worksheets[0].column_dimensions['T'].width = 10.7109375
gg.worksheets[0].column_dimensions['U'].width = 9.140625
gg.worksheets[1].row_dimensions[1].height = 45.0